from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from matplotlib.ticker import MaxNLocator


WORKBOOK = Path("data/subjects_info/AllDressed_PD_Participant_Study_Visit_Info.xlsx")
OUTPUT_DIR = Path("figures/minutes_since_last_dose")
SUBJECTS = [
    "PS_PD001",
    "PS_PD002",
    "PS_PD003",
    "PS_PD004",
    "PS_PD007",
    "PS_PD009",
    "PS_PD010",
    "PS_PD011",
    "PS_PD012",
    "PS_PD013",
    "PS_PD014",
    "PS_PD015",
    "PS_PD017",
    "PS_PD018",
    "PS_PD020",
    "PS_PD023",
    "PS_PD024",
    "PS_PD028",
]


def load_values(suffix: str) -> np.ndarray:
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    worksheet = workbook.active
    headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    dose_row = next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True))
    values_by_header = dict(zip(headers, dose_row))

    values = []
    for subject in SUBJECTS:
        header = f"{subject.replace('_', '')}_{suffix}"
        value = values_by_header.get(header)
        if not isinstance(value, (int, float)):
            raise ValueError(f"Missing numeric dose timing for {header}: {value!r}")
        values.append(value)
    return np.asarray(values, dtype=float)


def plot_histogram(values: np.ndarray, color: str, filename: str) -> None:
    median = float(np.median(values))

    fig, ax = plt.subplots(figsize=(7.2, 4.8))
    bin_edges = np.histogram_bin_edges(values, bins="auto")
    ax.hist(values, bins=bin_edges, color=color, edgecolor="white", linewidth=1.2)
    ax.axvline(median, color="#222222", linestyle="--", linewidth=1.8,
               label=f"Median = {median:g} min")

    ax.set_xlabel("Minutes since last dose")
    ax.set_ylabel("Number of participants")
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))
    ax.grid(axis="y", alpha=0.25)
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(OUTPUT_DIR / filename, dpi=300, bbox_inches="tight")
    fig.savefig(OUTPUT_DIR / Path(filename).with_suffix(".pdf"), bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    plot_histogram(load_values("OFF"), "#4472C4", "off_medication_histogram.png")
    plot_histogram(load_values("ON"), "#ED7D31", "on_medication_histogram.png")


if __name__ == "__main__":
    main()
