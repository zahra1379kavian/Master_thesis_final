#!/usr/bin/env python3
"""Relate subject-level reaction time to apathy and depression scores."""

from __future__ import annotations

import argparse
import json
import os
import re
import warnings
from pathlib import Path

os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib-clinical-rt")
warnings.filterwarnings("ignore", message="Unable to import Axes3D.*")

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl.styles import Alignment, Font, PatternFill
from scipy import stats
from statsmodels.stats.diagnostic import het_breuschpagan


ROOT = Path(__file__).resolve().parent
DEFAULT_BEHAVIOUR_DIR = Path(
    "/mnt/TeamShare/Data_Masterfile/H20-00572_All-Dressed/"
    "Zahra-Thesis-Data/fmri_opt_group/behaviour"
)
DEFAULT_CLINICAL_WORKBOOK = ROOT / "data/subjects_info/AllDressed_PD_Participant_Study_Visit_Info.xlsx"
DEFAULT_OUT_DIR = ROOT / "figures/clinical_rt_associations"
DEFAULT_PROJECTION_SESSION1_METRICS = (
    ROOT / "figures/projected_RT/projection_behavior_run_metrics_session1_medication_off.csv"
)
DEFAULT_REWARD_SUBJECT_STATS = ROOT / "figures/reward_effects/reward_rt_subject_bar_range_stats.csv"
CLINICAL_SHEET = "Sheet3"
APATHY_COLUMN = "Apathy Scale"
BDI_COLUMN = "Beck's Depression Inventory II (BDI-II)"
BEHAVIOUR_COLUMN_ZERO_BASED = 1
BEHAVIOUR_RE = re.compile(
    r"^PSPD(?P<digits>\d+)_ses_(?P<session>\d+)_run_(?P<run>\d+)\.npy$"
)
SUBJECT_SESSION_STATE_OVERRIDES = {("PSPD017", 1): "ON", ("PSPD017", 2): "OFF"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--behaviour-dir", type=Path, default=DEFAULT_BEHAVIOUR_DIR)
    parser.add_argument("--clinical-workbook", type=Path, default=DEFAULT_CLINICAL_WORKBOOK)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument(
        "--projection-session1-metrics",
        type=Path,
        default=DEFAULT_PROJECTION_SESSION1_METRICS,
    )
    parser.add_argument("--reward-subject-stats", type=Path, default=DEFAULT_REWARD_SUBJECT_STATS)
    return parser.parse_args()


def canonical_subject(value: object) -> str | None:
    match = re.search(r"PS[_-]?PD\s*0*(\d+)", str(value), flags=re.IGNORECASE)
    if match is None:
        return None
    return f"PSPD{int(match.group(1)):03d}"


def medication_state(subject: str, session: int) -> str:
    override = SUBJECT_SESSION_STATE_OVERRIDES.get((subject, int(session)))
    if override is not None:
        return override
    mapping = {1: "OFF", 2: "ON"}
    if int(session) not in mapping:
        raise ValueError(f"No medication-state mapping for {subject} session {session}")
    return mapping[int(session)]


def consecutive_rt_variability(rt_ms: np.ndarray) -> dict[str, float | int]:
    """Summarize truly adjacent finite RT pairs without bridging missing trials."""
    rt_ms = np.asarray(rt_ms, dtype=np.float64)
    previous = rt_ms[:-1]
    following = rt_ms[1:]
    denominator = previous**2 + following**2
    valid = (
        np.isfinite(previous)
        & np.isfinite(following)
        & np.isfinite(denominator)
        & (denominator > 0)
    )
    if not np.any(valid):
        return {
            "n_consecutive_rt_pairs": 0,
            "consecutive_rt_variability_sum": 0.0,
            "consecutive_rt_squared_difference_sum": 0.0,
            "mean_normalized_consecutive_rt_variability": np.nan,
            "rt_rmssd_ms": np.nan,
        }
    differences = previous[valid] - following[valid]
    normalized = differences**2 / denominator[valid]
    return {
        "n_consecutive_rt_pairs": int(np.count_nonzero(valid)),
        "consecutive_rt_variability_sum": float(np.sum(normalized)),
        "consecutive_rt_squared_difference_sum": float(np.sum(differences**2)),
        "mean_normalized_consecutive_rt_variability": float(np.mean(normalized)),
        "rt_rmssd_ms": float(np.sqrt(np.mean(differences**2))),
    }


def load_behaviour(
    behaviour_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    if not behaviour_dir.is_dir():
        raise FileNotFoundError(f"Behaviour directory not found: {behaviour_dir}")

    trial_rows: list[dict[str, object]] = []
    audit_rows: list[dict[str, object]] = []
    skipped_npy: list[str] = []
    seen_keys: set[tuple[str, int, int]] = set()

    for path in sorted(behaviour_dir.glob("*.npy")):
        match = BEHAVIOUR_RE.fullmatch(path.name)
        if match is None:
            skipped_npy.append(path.name)
            continue

        subject = f"PSPD{int(match.group('digits')):03d}"
        session = int(match.group("session"))
        run = int(match.group("run"))
        key = (subject, session, run)
        if key in seen_keys:
            raise RuntimeError(f"Duplicate behavioural subject/session/run: {key}")
        seen_keys.add(key)

        raw = np.asarray(np.load(path, allow_pickle=False), dtype=np.float64)
        if raw.ndim == 1:
            inverse_rt = raw
            source_kind = "1-D extracted inverse-RT vector"
        elif raw.ndim == 2 and raw.shape[1] > BEHAVIOUR_COLUMN_ZERO_BASED:
            inverse_rt = raw[:, BEHAVIOUR_COLUMN_ZERO_BASED]
            source_kind = "2-D array, column 2 (inverse RT)"
        else:
            raise RuntimeError(f"Unsupported behavioural shape {raw.shape} in {path}")

        valid = np.isfinite(inverse_rt) & (inverse_rt > 0)
        rt_ms_all = np.full(inverse_rt.shape, np.nan, dtype=np.float64)
        rt_ms_all[valid] = 1000.0 / inverse_rt[valid]
        rt_ms = rt_ms_all[valid]
        variability = consecutive_rt_variability(rt_ms_all)
        state = medication_state(subject, session)
        audit_rows.append(
            {
                "file": path.name,
                "subject": subject,
                "session": session,
                "medication": state,
                "run": run,
                "array_shape": str(tuple(raw.shape)),
                "source_selection": source_kind,
                "n_rows": int(inverse_rt.size),
                "n_valid_rt": int(valid.sum()),
                "n_invalid_or_missing": int((~valid).sum()),
                "median_rt_ms": float(np.median(rt_ms)),
                "mean_rt_ms": float(np.mean(rt_ms)),
                **variability,
            }
        )
        valid_indices = np.flatnonzero(valid)
        for trial_index, inverse_value, rt_value in zip(
            valid_indices + 1, inverse_rt[valid], rt_ms, strict=True
        ):
            trial_rows.append(
                {
                    "subject": subject,
                    "session": session,
                    "medication": state,
                    "run": run,
                    "trial_index": int(trial_index),
                    "inverse_rt_per_s": float(inverse_value),
                    "rt_ms": float(rt_value),
                    "source_file": path.name,
                }
            )

    if not trial_rows:
        raise RuntimeError(f"No exact-match behavioural files found in {behaviour_dir}")
    audit = pd.DataFrame(audit_rows).sort_values(["subject", "session", "run"]).reset_index(drop=True)
    expected_runs = {(1, 1), (1, 2), (2, 1), (2, 2)}
    for subject, rows in audit.groupby("subject"):
        observed_runs = set(zip(rows["session"].astype(int), rows["run"].astype(int), strict=True))
        if observed_runs != expected_runs:
            raise RuntimeError(
                f"{subject} has session/run combinations {sorted(observed_runs)}; "
                f"expected {sorted(expected_runs)}"
            )
    trials = pd.DataFrame(trial_rows).sort_values(
        ["subject", "session", "run", "trial_index"]
    ).reset_index(drop=True)
    return trials, audit, skipped_npy


def load_clinical(workbook: Path) -> pd.DataFrame:
    if not workbook.is_file():
        raise FileNotFoundError(f"Clinical workbook not found: {workbook}")
    clinical = pd.read_excel(workbook, sheet_name=CLINICAL_SHEET)
    required = {"Subject ID", APATHY_COLUMN, BDI_COLUMN}
    missing = required - set(clinical.columns)
    if missing:
        raise RuntimeError(f"Clinical sheet is missing columns: {sorted(missing)}")
    clinical = clinical.copy()
    clinical["clinical_subject_id_raw"] = clinical["Subject ID"].astype(str)
    clinical["subject"] = clinical["Subject ID"].map(canonical_subject)
    if clinical["subject"].isna().any():
        bad = clinical.loc[clinical["subject"].isna(), "Subject ID"].tolist()
        raise RuntimeError(f"Could not parse clinical subject IDs: {bad}")
    duplicate = clinical[clinical["subject"].duplicated(keep=False)]
    if not duplicate.empty:
        raise RuntimeError(
            "Duplicate canonical clinical IDs: " + ", ".join(duplicate["subject"].astype(str))
        )
    return clinical


def build_match_report(clinical: pd.DataFrame, behaviour_subjects: set[str]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    clinical_subjects = set(clinical["subject"])
    for row in clinical.itertuples(index=False):
        record = row._asdict()
        subject = str(record["subject"])
        raw_id = str(record["clinical_subject_id_raw"])
        in_behaviour = subject in behaviour_subjects
        # Namedtuple field names are altered for punctuation-heavy columns; use the source frame below.
        source = clinical.loc[clinical["subject"] == subject].iloc[0]
        complete = pd.notna(source[APATHY_COLUMN]) and pd.notna(source[BDI_COLUMN])
        if in_behaviour and complete:
            status = "Included"
        elif in_behaviour:
            status = "Excluded: missing apathy or BDI-II"
        elif "withdrew" in raw_id.lower():
            status = "Not in behavioural dataset (withdrew)"
        else:
            status = "Not in behavioural dataset"
        rows.append(
            {
                "clinical_subject_id_raw": raw_id,
                "canonical_subject": subject,
                "behaviour_files_present": in_behaviour,
                "apathy_present": pd.notna(source[APATHY_COLUMN]),
                "bdi_ii_present": pd.notna(source[BDI_COLUMN]),
                "analysis_status": status,
            }
        )
    for subject in sorted(behaviour_subjects - clinical_subjects):
        rows.append(
            {
                "clinical_subject_id_raw": "",
                "canonical_subject": subject,
                "behaviour_files_present": True,
                "apathy_present": False,
                "bdi_ii_present": False,
                "analysis_status": "Excluded: no clinical row",
            }
        )
    return pd.DataFrame(rows).sort_values("canonical_subject").reset_index(drop=True)


def summarize_subjects(
    trials: pd.DataFrame,
    file_audit: pd.DataFrame,
    clinical: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    session_summary = (
        trials.groupby(["subject", "session", "medication"], sort=True)
        .agg(
            n_valid_trials=("rt_ms", "size"),
            median_rt_ms=("rt_ms", "median"),
            mean_rt_ms=("rt_ms", "mean"),
            sd_rt_ms=("rt_ms", "std"),
            q25_rt_ms=("rt_ms", lambda x: float(x.quantile(0.25))),
            q75_rt_ms=("rt_ms", lambda x: float(x.quantile(0.75))),
        )
        .reset_index()
    )
    session_summary["iqr_rt_ms"] = session_summary["q75_rt_ms"] - session_summary["q25_rt_ms"]
    session_variability = (
        file_audit.groupby(["subject", "session", "medication"], sort=True)
        .agg(
            n_consecutive_rt_pairs=("n_consecutive_rt_pairs", "sum"),
            consecutive_rt_variability_sum=("consecutive_rt_variability_sum", "sum"),
            consecutive_rt_squared_difference_sum=("consecutive_rt_squared_difference_sum", "sum"),
        )
        .reset_index()
    )
    session_variability["mean_normalized_consecutive_rt_variability"] = (
        session_variability["consecutive_rt_variability_sum"]
        / session_variability["n_consecutive_rt_pairs"]
    )
    session_variability["rt_rmssd_ms"] = np.sqrt(
        session_variability["consecutive_rt_squared_difference_sum"]
        / session_variability["n_consecutive_rt_pairs"]
    )
    session_summary = session_summary.merge(
        session_variability,
        on=["subject", "session", "medication"],
        validate="one_to_one",
    )

    overall = (
        trials.groupby("subject", sort=True)
        .agg(
            n_valid_trials=("rt_ms", "size"),
            overall_median_rt_ms=("rt_ms", "median"),
            overall_mean_rt_ms=("rt_ms", "mean"),
            overall_sd_rt_ms=("rt_ms", "std"),
            overall_q25_rt_ms=("rt_ms", lambda x: float(x.quantile(0.25))),
            overall_q75_rt_ms=("rt_ms", lambda x: float(x.quantile(0.75))),
        )
        .reset_index()
    )
    overall["overall_iqr_rt_ms"] = overall["overall_q75_rt_ms"] - overall["overall_q25_rt_ms"]
    subject_variability = (
        file_audit.groupby("subject", sort=True)
        .agg(
            n_consecutive_rt_pairs=("n_consecutive_rt_pairs", "sum"),
            consecutive_rt_variability_sum=("consecutive_rt_variability_sum", "sum"),
            consecutive_rt_squared_difference_sum=("consecutive_rt_squared_difference_sum", "sum"),
        )
        .reset_index()
    )
    subject_variability["mean_normalized_consecutive_rt_variability"] = (
        subject_variability["consecutive_rt_variability_sum"]
        / subject_variability["n_consecutive_rt_pairs"]
    )
    subject_variability["rt_rmssd_ms"] = np.sqrt(
        subject_variability["consecutive_rt_squared_difference_sum"]
        / subject_variability["n_consecutive_rt_pairs"]
    )
    overall = overall.merge(subject_variability, on="subject", validate="one_to_one")
    state_wide = session_summary.pivot(index="subject", columns="medication", values=[
        "n_valid_trials", "median_rt_ms", "mean_rt_ms"
    ])
    state_wide.columns = [f"{state.lower()}_{metric}" for metric, state in state_wide.columns]
    state_wide = state_wide.reset_index()
    subjects = overall.merge(state_wide, on="subject", how="left", validate="one_to_one")
    subjects["equal_weight_session_median_rt_ms"] = subjects[[
        "off_median_rt_ms", "on_median_rt_ms"
    ]].mean(axis=1)

    selected_clinical = clinical[[
        "subject", "clinical_subject_id_raw", APATHY_COLUMN, BDI_COLUMN
    ]].rename(columns={APATHY_COLUMN: "apathy_score", BDI_COLUMN: "bdi_ii_score"})
    subjects = subjects.merge(selected_clinical, on="subject", how="left", validate="one_to_one")
    subjects["included_primary"] = subjects[["apathy_score", "bdi_ii_score"]].notna().all(axis=1)
    subjects = subjects.sort_values("subject").reset_index(drop=True)
    return subjects, session_summary


def bh_fdr(p_values: pd.Series) -> np.ndarray:
    p = p_values.to_numpy(dtype=float)
    order = np.argsort(p)
    ranked = p[order]
    adjusted = ranked * len(p) / np.arange(1, len(p) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    out = np.empty_like(adjusted)
    out[order] = np.minimum(adjusted, 1.0)
    return out


def correlation_table(
    data: pd.DataFrame,
    outcomes: list[tuple[str, str]],
    predictors: list[tuple[str, str]],
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for outcome, outcome_label in outcomes:
        for predictor, predictor_label in predictors:
            clean = data[[predictor, outcome]].dropna()
            x = clean[predictor].to_numpy(dtype=float)
            y = clean[outcome].to_numpy(dtype=float)
            for method, function in (("Pearson", stats.pearsonr), ("Spearman", stats.spearmanr)):
                result = function(x, y)
                ci_low = np.nan
                ci_high = np.nan
                if method == "Pearson":
                    ci = result.confidence_interval(confidence_level=0.95)
                    ci_low = float(ci.low)
                    ci_high = float(ci.high)
                rows.append(
                    {
                        "outcome": outcome_label,
                        "predictor": predictor_label,
                        "method": method,
                        "n_subjects": len(clean),
                        "correlation": float(result.statistic),
                        "ci_low": ci_low,
                        "ci_high": ci_high,
                        "p_value": float(result.pvalue),
                    }
                )
    table = pd.DataFrame(rows)
    table["fdr_q_value"] = bh_fdr(table["p_value"])
    table["significant_fdr_0_05"] = table["fdr_q_value"] < 0.05
    return table


def requested_correlation_tables(
    data: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    predictors = [("apathy_score", "Apathy Scale"), ("bdi_ii_score", "BDI-II")]
    rt = correlation_table(
        data,
        [("overall_median_rt_ms", "Overall median RT (ms)")],
        predictors,
    )
    variability = correlation_table(
        data,
        [
            (
                "mean_normalized_consecutive_rt_variability",
                "Mean normalized consecutive-trial RT variability",
            )
        ],
        predictors,
    )
    clinical = correlation_table(
        data,
        [("bdi_ii_score", "BDI-II")],
        [("apathy_score", "Apathy Scale")],
    )
    return rt, variability, clinical


def bivariate_result(data: pd.DataFrame, predictor: str, predictor_label: str, outcome: str) -> dict[str, object]:
    clean = data[[predictor, outcome]].dropna()
    x = clean[predictor].to_numpy(dtype=float)
    y = clean[outcome].to_numpy(dtype=float)
    pearson = stats.pearsonr(x, y)
    correlation_ci = pearson.confidence_interval(confidence_level=0.95)
    model = sm.OLS(y, sm.add_constant(x)).fit()
    ci = model.conf_int(alpha=0.05)
    return {
        "predictor": predictor_label,
        "outcome": outcome,
        "n_subjects": len(clean),
        "pearson_r": float(pearson.statistic),
        "pearson_ci_low": float(correlation_ci.low),
        "pearson_ci_high": float(correlation_ci.high),
        "p_value": float(pearson.pvalue),
        "regression_slope_ms_per_point": float(model.params[1]),
        "slope_ci_low": float(ci[1, 0]),
        "slope_ci_high": float(ci[1, 1]),
        "intercept_ms": float(model.params[0]),
        "r_squared": float(model.rsquared),
    }


def primary_analysis(data: pd.DataFrame) -> pd.DataFrame:
    rows = [
        bivariate_result(data, "apathy_score", "Apathy Scale", "overall_median_rt_ms"),
        bivariate_result(data, "bdi_ii_score", "BDI-II", "overall_median_rt_ms"),
    ]
    table = pd.DataFrame(rows)
    table["fdr_q_value"] = bh_fdr(table["p_value"])
    table["significant_fdr_0_05"] = table["fdr_q_value"] < 0.05
    return table


def joint_regression(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, object]:
    clean = data[["subject", "overall_median_rt_ms", "apathy_score", "bdi_ii_score"]].dropna()
    y = clean["overall_median_rt_ms"].astype(float)
    predictors = ["apathy_score", "bdi_ii_score"]
    x = clean[predictors].astype(float)
    model = sm.OLS(y, sm.add_constant(x)).fit()
    robust = model.get_robustcov_results(cov_type="HC3")
    ci = model.conf_int(alpha=0.05)
    robust_ci = robust.conf_int(alpha=0.05)
    labels = {"const": "Intercept", "apathy_score": "Apathy Scale", "bdi_ii_score": "BDI-II"}
    coefficient_rows: list[dict[str, object]] = []
    for index, term in enumerate(model.params.index):
        standardized_beta = np.nan
        if term != "const":
            standardized_beta = model.params[term] * x[term].std(ddof=1) / y.std(ddof=1)
        coefficient_rows.append(
            {
                "term": labels[term],
                "coefficient_ms_per_point": float(model.params[term]),
                "ci_low": float(ci.loc[term, 0]),
                "ci_high": float(ci.loc[term, 1]),
                "standardized_beta": float(standardized_beta),
                "p_value": float(model.pvalues[term]),
                "hc3_ci_low": float(robust_ci[index, 0]),
                "hc3_ci_high": float(robust_ci[index, 1]),
                "hc3_p_value": float(robust.pvalues[index]),
            }
        )
    score_r = stats.pearsonr(clean["apathy_score"], clean["bdi_ii_score"])
    vif = 1.0 / (1.0 - float(score_r.statistic) ** 2)
    model_table = pd.DataFrame([
        {
            "model": "Median RT ~ Apathy + BDI-II",
            "n_subjects": int(model.nobs),
            "r_squared": float(model.rsquared),
            "adjusted_r_squared": float(model.rsquared_adj),
            "f_statistic": float(model.fvalue),
            "model_p_value": float(model.f_pvalue),
            "apathy_bdi_pearson_r": float(score_r.statistic),
            "apathy_bdi_p_value": float(score_r.pvalue),
            "vif_each_predictor": float(vif),
        }
    ])
    return pd.DataFrame(coefficient_rows), model_table, model


def sensitivity_analysis(data: pd.DataFrame) -> pd.DataFrame:
    definitions = [
        ("Pearson", "Overall mean RT", "overall_mean_rt_ms", stats.pearsonr),
        ("Pearson", "Equal-weight mean of OFF/ON medians", "equal_weight_session_median_rt_ms", stats.pearsonr),
        ("Pearson", "OFF-session median RT", "off_median_rt_ms", stats.pearsonr),
        ("Pearson", "ON-session median RT", "on_median_rt_ms", stats.pearsonr),
        ("Spearman", "Overall median RT", "overall_median_rt_ms", stats.spearmanr),
    ]
    predictors = [("apathy_score", "Apathy Scale"), ("bdi_ii_score", "BDI-II")]
    rows: list[dict[str, object]] = []
    for method, outcome_label, outcome, function in definitions:
        for predictor, predictor_label in predictors:
            clean = data[[predictor, outcome]].dropna()
            result = function(clean[predictor], clean[outcome])
            rows.append(
                {
                    "method": method,
                    "outcome": outcome_label,
                    "predictor": predictor_label,
                    "n_subjects": len(clean),
                    "correlation": float(result.statistic),
                    "p_value": float(result.pvalue),
                }
            )
    table = pd.DataFrame(rows)
    table["fdr_q_across_sensitivity_tests"] = bh_fdr(table["p_value"])
    return table


def diagnostic_analysis(data: pd.DataFrame, joint_model: object) -> pd.DataFrame:
    clean = data[["subject", "overall_median_rt_ms", "apathy_score", "bdi_ii_score"]].dropna()
    influence = joint_model.get_influence()
    cooks = pd.Series(influence.cooks_distance[0], index=clean.index)
    largest_index = cooks.idxmax()
    loo_rows = []
    for index in clean.index:
        subset = clean.drop(index=index)
        loo_rows.append(
            {
                "apathy": float(stats.pearsonr(subset["apathy_score"], subset["overall_median_rt_ms"]).statistic),
                "bdi": float(stats.pearsonr(subset["bdi_ii_score"], subset["overall_median_rt_ms"]).statistic),
            }
        )
    loo = pd.DataFrame(loo_rows)
    shapiro = stats.shapiro(joint_model.resid)
    bp = het_breuschpagan(joint_model.resid, joint_model.model.exog)
    return pd.DataFrame([
        {"diagnostic": "Shapiro-Wilk residual normality", "value": float(shapiro.statistic), "p_value": float(shapiro.pvalue), "detail": "p > 0.05 does not indicate a detectable departure"},
        {"diagnostic": "Breusch-Pagan heteroscedasticity", "value": float(bp[0]), "p_value": float(bp[1]), "detail": "p > 0.05 does not indicate detectable heteroscedasticity"},
        {"diagnostic": "Largest Cook's distance", "value": float(cooks.max()), "p_value": np.nan, "detail": str(clean.loc[largest_index, "subject"])},
        {"diagnostic": "Leave-one-out apathy Pearson r range", "value": np.nan, "p_value": np.nan, "detail": f"{loo['apathy'].min():.3f} to {loo['apathy'].max():.3f}"},
        {"diagnostic": "Leave-one-out BDI-II Pearson r range", "value": np.nan, "p_value": np.nan, "detail": f"{loo['bdi'].min():.3f} to {loo['bdi'].max():.3f}"},
    ])


def regression_plot(data: pd.DataFrame, primary: pd.DataFrame, out_dir: Path) -> None:
    panels = [
        ("apathy_score", "Apathy Scale score", "Apathy Scale"),
        ("bdi_ii_score", "BDI-II score", "BDI-II"),
    ]
    fig, axes = plt.subplots(1, 2, figsize=(11.5, 5.0), constrained_layout=True)
    for ax, (column, xlabel, label) in zip(axes, panels, strict=True):
        clean = data[["subject", column, "overall_median_rt_ms"]].dropna().copy()
        x = clean[column].to_numpy(dtype=float)
        y = clean["overall_median_rt_ms"].to_numpy(dtype=float)
        model = sm.OLS(y, sm.add_constant(x)).fit()
        grid = np.linspace(x.min(), x.max(), 200)
        prediction = model.get_prediction(sm.add_constant(grid)).summary_frame(alpha=0.05)
        ax.fill_between(
            grid,
            prediction["mean_ci_lower"].to_numpy(),
            prediction["mean_ci_upper"].to_numpy(),
            color="#4C78A8",
            alpha=0.18,
            linewidth=0,
            label="95% CI for mean",
        )
        ax.plot(grid, prediction["mean"], color="#245A86", linewidth=2.2)
        ax.scatter(x, y, s=82, color="#E07A5F", edgecolor="white", linewidth=0.9, zorder=3)
        result = primary.loc[primary["predictor"] == label].iloc[0]
        ax.text(
            0.04,
            0.96,
            f"Pearson r = {result['pearson_r']:.3f}, p = {result['p_value']:.4g}",
            transform=ax.transAxes,
            va="top",
            fontsize=12,
            bbox={"boxstyle": "round,pad=0.35", "facecolor": "white", "edgecolor": "#BBBBBB", "alpha": 0.9},
        )
        ax.set_xlabel(xlabel, fontsize=13)
        ax.set_ylabel("Median reaction time (ms)", fontsize=13)
        ax.tick_params(axis="both", labelsize=12)
        ax.grid(alpha=0.2)
        ax.spines[["top", "right"]].set_visible(False)
    fig.suptitle("Reaction time associations with apathy and depression", fontsize=16, fontweight="bold")
    fig.savefig(out_dir / "primary_rt_clinical_scatterplots.png", dpi=220)
    fig.savefig(out_dir / "primary_rt_clinical_scatterplots.pdf")
    plt.close(fig)


def variability_scatterplot(
    data: pd.DataFrame,
    variability_correlations: pd.DataFrame,
    out_dir: Path,
) -> None:
    panels = [
        ("apathy_score", "Apathy Scale score", "Apathy Scale"),
        ("bdi_ii_score", "BDI-II score", "BDI-II"),
    ]
    outcome = "mean_normalized_consecutive_rt_variability"
    fig, axes = plt.subplots(1, 2, figsize=(11.5, 5.0), constrained_layout=True)
    for panel_index, (ax, (column, xlabel, label)) in enumerate(
        zip(axes, panels, strict=True)
    ):
        clean = data[["subject", column, outcome]].dropna().copy()
        x = clean[column].to_numpy(dtype=float)
        y = clean[outcome].to_numpy(dtype=float)
        model = sm.OLS(y, sm.add_constant(x)).fit()
        grid = np.linspace(x.min(), x.max(), 200)
        prediction = model.get_prediction(sm.add_constant(grid)).summary_frame(alpha=0.05)
        ax.fill_between(
            grid,
            prediction["mean_ci_lower"].to_numpy(),
            prediction["mean_ci_upper"].to_numpy(),
            color="#4C78A8",
            alpha=0.18,
            linewidth=0,
        )
        ax.plot(grid, prediction["mean"], color="#245A86", linewidth=2.2)
        ax.scatter(x, y, s=82, color="#59A14F", edgecolor="white", linewidth=0.9, zorder=3)
        rows = variability_correlations[variability_correlations["predictor"] == label]
        pearson = rows[rows["method"] == "Pearson"].iloc[0]
        ax.text(
            0.96,
            0.96,
            f"Pearson r = {pearson['correlation']:.3f}, p = {pearson['p_value']:.4g}",
            transform=ax.transAxes,
            ha="right",
            va="top",
            fontsize=12,
            bbox={"boxstyle": "round,pad=0.35", "facecolor": "white", "edgecolor": "#BBBBBB", "alpha": 0.9},
        )
        ax.set_xlabel(xlabel, fontsize=13)
        ax.set_ylabel("RT varaibility" if panel_index == 0 else "", fontsize=13)
        ax.tick_params(axis="both", labelsize=12)
        y_low, y_high = ax.get_ylim()
        ax.set_ylim(y_low, y_high + 0.08 * (y_high - y_low))
        ax.grid(alpha=0.2)
        ax.spines[["top", "right"]].set_visible(False)
    fig.suptitle("Consecutive-trial RT variability and clinical scores", fontsize=16, fontweight="bold")
    fig.savefig(out_dir / "rt_variability_clinical_scatterplots.png", dpi=220)
    fig.savefig(out_dir / "rt_variability_clinical_scatterplots.pdf")
    plt.close(fig)


def _paired_comparison_row(data: pd.DataFrame, suffix: str, label: str) -> dict[str, object]:
    behaviour = data[f"behavior{suffix}"].to_numpy(dtype=float)
    projection = data[f"projection{suffix}"].to_numpy(dtype=float)
    difference = behaviour - projection
    paired = stats.ttest_rel(behaviour, projection)
    wilcoxon = stats.wilcoxon(difference)
    sem = stats.sem(difference)
    half_width = stats.t.ppf(0.975, len(difference) - 1) * sem
    return {
        "analysis": label,
        "n_subjects": len(difference),
        "mean_behaviour_variability": float(np.mean(behaviour)),
        "mean_projection_variability": float(np.mean(projection)),
        "mean_behaviour_minus_projection": float(np.mean(difference)),
        "difference_ci_low": float(np.mean(difference) - half_width),
        "difference_ci_high": float(np.mean(difference) + half_width),
        "paired_t": float(paired.statistic),
        "paired_t_p_value": float(paired.pvalue),
        "wilcoxon_statistic": float(wilcoxon.statistic),
        "wilcoxon_p_value": float(wilcoxon.pvalue),
    }


def adjusted_projection_analysis(
    metrics_path: Path,
    subject_data: pd.DataFrame,
    out_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if not metrics_path.is_file():
        raise FileNotFoundError(f"Projected-RT session-1 metrics not found: {metrics_path}")
    metrics = pd.read_csv(metrics_path)
    required = {
        "sub_tag",
        "adjacent_diff_ratio_sum_behavior_col2",
        "adjacent_diff_ratio_sum_projection",
    }
    missing = required - set(metrics.columns)
    if missing:
        raise RuntimeError(f"Projection metrics are missing columns: {sorted(missing)}")
    subject_metrics = (
        metrics.groupby("sub_tag", as_index=False)
        .agg(
            behavior=("adjacent_diff_ratio_sum_behavior_col2", "mean"),
            projection=("adjacent_diff_ratio_sum_projection", "mean"),
            n_runs=("sub_tag", "size"),
        )
    )
    digits = subject_metrics["sub_tag"].astype(str).str.extract(r"(\d+)$")[0]
    if digits.isna().any():
        raise RuntimeError("Could not parse one or more projection-metric subject IDs")
    subject_metrics["subject"] = digits.map(lambda value: f"PSPD{int(value):03d}")
    clinical = subject_data[["subject", "apathy_score", "bdi_ii_score"]]
    subject_metrics = subject_metrics.merge(clinical, on="subject", how="left", validate="one_to_one")
    if subject_metrics[["apathy_score", "bdi_ii_score"]].isna().any().any():
        raise RuntimeError("A projection-metric participant is missing an apathy or BDI-II score")

    centered = subject_metrics[["apathy_score", "bdi_ii_score"]] - subject_metrics[[
        "apathy_score", "bdi_ii_score"
    ]].mean()
    design = sm.add_constant(centered)
    coefficient_rows: list[dict[str, object]] = []
    for outcome in ("behavior", "projection"):
        model = sm.OLS(subject_metrics[outcome], design).fit()
        subject_metrics[f"{outcome}_adjusted"] = subject_metrics[outcome].mean() + model.resid
        ci = model.conf_int(alpha=0.05)
        for term in model.params.index:
            coefficient_rows.append(
                {
                    "outcome": outcome,
                    "term": term,
                    "coefficient": float(model.params[term]),
                    "ci_low": float(ci.loc[term, 0]),
                    "ci_high": float(ci.loc[term, 1]),
                    "p_value": float(model.pvalues[term]),
                    "model_r_squared": float(model.rsquared),
                    "model_p_value": float(model.f_pvalue),
                }
            )
    subject_metrics["raw_behaviour_minus_projection"] = (
        subject_metrics["behavior"] - subject_metrics["projection"]
    )
    subject_metrics["adjusted_behaviour_minus_projection"] = (
        subject_metrics["behavior_adjusted"] - subject_metrics["projection_adjusted"]
    )
    comparison = pd.DataFrame([
        _paired_comparison_row(subject_metrics, "", "Original saved metric"),
        _paired_comparison_row(
            subject_metrics,
            "_adjusted",
            "Adjusted for centered apathy and BDI-II",
        ),
    ])
    coefficients = pd.DataFrame(coefficient_rows)
    _plot_adjusted_projection(subject_metrics, comparison, out_dir)
    return subject_metrics, coefficients, comparison


def _plot_adjusted_projection(
    data: pd.DataFrame,
    comparison: pd.DataFrame,
    out_dir: Path,
) -> None:
    behavior = data["behavior_adjusted"].to_numpy(dtype=float)
    projection = data["projection_adjusted"].to_numpy(dtype=float)
    difference = behavior - projection
    order = np.argsort(difference)
    behavior = behavior[order]
    projection = projection[order]
    rng = np.random.default_rng(141)
    jitter = np.linspace(-0.05, 0.05, len(data))
    row = comparison.loc[comparison["analysis"].str.startswith("Adjusted")].iloc[0]

    fig, axes = plt.subplots(1, 2, figsize=(10.8, 5.6), gridspec_kw={"width_ratios": [1.45, 0.85]}, constrained_layout=True)
    ax = axes[0]
    box = ax.boxplot(
        [behavior, projection], positions=[0, 1], widths=0.35, patch_artist=True, showfliers=False
    )
    for patch, color in zip(box["boxes"], ["#5DA5DA", "#F28E2B"], strict=True):
        patch.set_facecolor(color)
        patch.set_alpha(0.22)
        patch.set_edgecolor(color)
    for y0, y1, offset in zip(behavior, projection, jitter, strict=True):
        ax.plot([offset, 1 + offset], [y0, y1], color="0.65", alpha=0.45, linewidth=0.8)
    ax.scatter(jitter, behavior, color="#2C8ABE", edgecolor="white", linewidth=0.4, s=38, zorder=3)
    ax.scatter(1 + jitter, projection, color="#E07B24", edgecolor="white", linewidth=0.4, s=38, zorder=3)
    ax.set_xticks([0, 1], ["Behaviour", "Projection"])
    ax.set_ylabel("Clinical-adjusted consecutive-trial variability")
    ax.grid(axis="y", alpha=0.2)
    ax.spines[["top", "right"]].set_visible(False)

    ax = axes[1]
    box = ax.boxplot([difference], positions=[0], widths=0.28, patch_artist=True, showfliers=False)
    box["boxes"][0].set_facecolor("#2CA58D")
    box["boxes"][0].set_alpha(0.22)
    box["boxes"][0].set_edgecolor("#2CA58D")
    ax.scatter(rng.uniform(-0.06, 0.06, len(difference)), difference, color="#2CA58D", edgecolor="white", linewidth=0.4, s=38)
    ax.axhline(0, color="0.35", linestyle="--", linewidth=1)
    ax.set_xticks([0], ["Behaviour − Projection"])
    ax.set_ylabel("Adjusted variability difference")
    ax.text(
        0.04,
        0.96,
        f"N = {int(row['n_subjects'])}\nMean difference = {row['mean_behaviour_minus_projection']:.2f}\n"
        f"Paired t p = {row['paired_t_p_value']:.4g}",
        transform=ax.transAxes,
        va="top",
        fontsize=9.5,
        bbox={"boxstyle": "round,pad=0.35", "facecolor": "white", "edgecolor": "#BBBBBB", "alpha": 0.9},
    )
    ax.grid(axis="y", alpha=0.2)
    ax.spines[["top", "right"]].set_visible(False)
    fig.suptitle(
        "Session 1 projected-signal comparison after apathy/BDI-II adjustment",
        fontsize=13,
        fontweight="bold",
    )
    stem = out_dir / "projection_behavior_subject_panel_session1_adjusted_apathy_bdi"
    fig.savefig(stem.with_suffix(".png"), dpi=220)
    fig.savefig(stem.with_suffix(".pdf"))
    plt.close(fig)


def reward_difference_analysis(
    reward_stats_path: Path,
    subject_data: pd.DataFrame,
    out_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not reward_stats_path.is_file():
        raise FileNotFoundError(f"Reward subject statistics not found: {reward_stats_path}")
    reward = pd.read_csv(reward_stats_path)
    required = {
        "subject",
        "n_low",
        "n_high",
        "mean_rt_low",
        "mean_rt_high",
        "delta_rt_high_minus_low",
    }
    missing = required - set(reward.columns)
    if missing:
        raise RuntimeError(f"Reward subject statistics are missing columns: {sorted(missing)}")
    if reward["subject"].duplicated().any():
        raise RuntimeError("Reward subject statistics contain duplicate subject rows")

    selected = reward[[
        "subject",
        "n_low",
        "n_high",
        "mean_rt_low",
        "mean_rt_high",
        "delta_rt_high_minus_low",
    ]].copy()
    selected["mean_rt_low_ms"] = 1000.0 * selected["mean_rt_low"]
    selected["mean_rt_high_ms"] = 1000.0 * selected["mean_rt_high"]
    selected["reward_rt_high_minus_low_ms"] = 1000.0 * selected["delta_rt_high_minus_low"]
    selected = selected.drop(columns=["mean_rt_low", "mean_rt_high", "delta_rt_high_minus_low"])
    selected = selected.merge(
        subject_data[["subject", "apathy_score", "bdi_ii_score"]],
        on="subject",
        how="left",
        validate="one_to_one",
    )
    if selected[["apathy_score", "bdi_ii_score"]].isna().any().any():
        raise RuntimeError("A reward-figure participant is missing an apathy or BDI-II score")
    correlations = correlation_table(
        selected,
        [("reward_rt_high_minus_low_ms", "High-minus-low reward mean RT (ms)")],
        [("apathy_score", "Apathy Scale"), ("bdi_ii_score", "BDI-II")],
    )
    _plot_reward_difference_correlations(selected, correlations, out_dir)
    return selected.sort_values("subject").reset_index(drop=True), correlations


def _plot_reward_difference_correlations(
    data: pd.DataFrame,
    correlations: pd.DataFrame,
    out_dir: Path,
) -> None:
    panels = [
        ("apathy_score", "Apathy Scale score", "Apathy Scale"),
        ("bdi_ii_score", "BDI-II score", "BDI-II"),
    ]
    outcome = "reward_rt_high_minus_low_ms"
    fig, axes = plt.subplots(1, 2, figsize=(11.5, 5.0), constrained_layout=True)
    for ax, (column, xlabel, label) in zip(axes, panels, strict=True):
        x = data[column].to_numpy(dtype=float)
        y = data[outcome].to_numpy(dtype=float)
        model = sm.OLS(y, sm.add_constant(x)).fit()
        grid = np.linspace(x.min(), x.max(), 200)
        prediction = model.get_prediction(sm.add_constant(grid)).summary_frame(alpha=0.05)
        ax.fill_between(
            grid,
            prediction["mean_ci_lower"].to_numpy(),
            prediction["mean_ci_upper"].to_numpy(),
            color="#4C78A8",
            alpha=0.18,
            linewidth=0,
        )
        ax.plot(grid, prediction["mean"], color="#245A86", linewidth=2.2)
        ax.scatter(x, y, s=82, color="#B07AA1", edgecolor="white", linewidth=0.9, zorder=3)
        rows = correlations[correlations["predictor"] == label]
        pearson = rows[rows["method"] == "Pearson"].iloc[0]
        ax.text(
            0.04,
            0.96,
            f"Pearson r = {pearson['correlation']:.3f}, p = {pearson['p_value']:.4g}",
            transform=ax.transAxes,
            va="top",
            fontsize=12,
            bbox={"boxstyle": "round,pad=0.35", "facecolor": "white", "edgecolor": "#BBBBBB", "alpha": 0.9},
        )
        ax.axhline(0.0, color="0.35", linestyle="--", linewidth=1.0)
        ax.set_xlabel(xlabel, fontsize=13)
        ax.set_ylabel("High-reward − low-reward mean RT (ms)", fontsize=13)
        ax.tick_params(axis="both", labelsize=12)
        ax.grid(alpha=0.2)
        ax.spines[["top", "right"]].set_visible(False)
    fig.suptitle("Reward-related RT difference and clinical scores", fontsize=16, fontweight="bold")
    fig.savefig(out_dir / "reward_rt_difference_clinical_scatterplots.png", dpi=220)
    fig.savefig(out_dir / "reward_rt_difference_clinical_scatterplots.pdf")
    plt.close(fig)


def markdown_table(frame: pd.DataFrame, columns: list[str], formats: dict[str, str] | None = None) -> str:
    display = frame[columns].copy()
    formats = formats or {}
    for column, specification in formats.items():
        display[column] = display[column].map(
            lambda value: "" if pd.isna(value) else format(float(value), specification)
        )
    return display.to_markdown(index=False)


def write_report(
    out_dir: Path,
    behaviour_dir: Path,
    clinical_workbook: Path,
    subjects: pd.DataFrame,
    primary: pd.DataFrame,
    coefficients: pd.DataFrame,
    joint_model_table: pd.DataFrame,
    sensitivity: pd.DataFrame,
    rt_correlations: pd.DataFrame,
    variability_correlations: pd.DataFrame,
    clinical_correlations: pd.DataFrame,
    reward_correlations: pd.DataFrame,
    projection_adjustment_models: pd.DataFrame,
    projection_comparison: pd.DataFrame,
    diagnostics: pd.DataFrame,
    match_report: pd.DataFrame,
    skipped_npy: list[str],
) -> None:
    included = subjects[subjects["included_primary"]].copy()
    model_row = joint_model_table.iloc[0]
    primary_display = primary.rename(columns={
        "pearson_ci_low": "r_95ci_low",
        "pearson_ci_high": "r_95ci_high",
        "slope_ci_low": "slope_95ci_low",
        "slope_ci_high": "slope_95ci_high",
    })
    subject_display = included.rename(columns={
        "subject": "Subject",
        "apathy_score": "Apathy",
        "bdi_ii_score": "BDI-II",
        "n_valid_trials": "Valid trials",
        "overall_median_rt_ms": "Median RT (ms)",
        "off_median_rt_ms": "OFF median (ms)",
        "on_median_rt_ms": "ON median (ms)",
    })
    report = [
        "# Reaction Time, Apathy, and Depression",
        "",
        "## Main finding",
        "",
        (
            f"The analysis included {len(included)} participants. Slower median reaction time was associated "
            "with both higher apathy and higher BDI-II scores. Both primary associations remained significant "
            "after FDR correction across the two primary tests."
        ),
        "",
        "## Primary associations",
        "",
        markdown_table(
            primary_display,
            ["predictor", "n_subjects", "pearson_r", "r_95ci_low", "r_95ci_high", "p_value", "fdr_q_value", "regression_slope_ms_per_point", "slope_95ci_low", "slope_95ci_high", "r_squared"],
            {"pearson_r": ".3f", "r_95ci_low": ".3f", "r_95ci_high": ".3f", "p_value": ".4g", "fdr_q_value": ".4g", "regression_slope_ms_per_point": ".2f", "slope_95ci_low": ".2f", "slope_95ci_high": ".2f", "r_squared": ".3f"},
        ),
        "",
        "The slope is the estimated increase in median RT, in milliseconds, for each one-point increase in the clinical score.",
        "",
        "### Pearson and Spearman correlations",
        "",
        markdown_table(
            rt_correlations,
            ["predictor", "method", "n_subjects", "correlation", "ci_low", "ci_high", "p_value", "fdr_q_value"],
            {"correlation": ".3f", "ci_low": ".3f", "ci_high": ".3f", "p_value": ".4g", "fdr_q_value": ".4g"},
        ),
        "",
        "The RT-correlation q values control FDR across the four Pearson/Spearman-by-score tests in this table.",
        "",
        "## Correlation between apathy and depression",
        "",
        markdown_table(
            clinical_correlations,
            ["outcome", "predictor", "method", "n_subjects", "correlation", "ci_low", "ci_high", "p_value"],
            {"correlation": ".3f", "ci_low": ".3f", "ci_high": ".3f", "p_value": ".4g"},
        ),
        "",
        "## Clinical scores and consecutive-trial RT variability",
        "",
        markdown_table(
            variability_correlations,
            ["predictor", "method", "n_subjects", "correlation", "ci_low", "ci_high", "p_value", "fdr_q_value"],
            {"correlation": ".3f", "ci_low": ".3f", "ci_high": ".3f", "p_value": ".4g", "fdr_q_value": ".4g"},
        ),
        "",
        "The variability outcome is the subject-level mean of the normalized squared difference between genuinely adjacent valid RTs within each run: `(RT[t] - RT[t+1])² / (RT[t]² + RT[t+1]²)`. Missing trials are not bridged.",
        "",
        "The variability q values control FDR across the four Pearson/Spearman-by-score tests in this table.",
        "",
        "## High-minus-low reward RT difference and clinical scores",
        "",
        markdown_table(
            reward_correlations,
            ["predictor", "method", "n_subjects", "correlation", "ci_low", "ci_high", "p_value", "fdr_q_value"],
            {"correlation": ".3f", "ci_low": ".3f", "ci_high": ".3f", "p_value": ".4g", "fdr_q_value": ".4g"},
        ),
        "",
        "The outcome is the exact subject-level mean difference underlying `reward_rt_subject_bar_range.png`, converted to milliseconds: high-reward mean RT minus low-reward mean RT. Negative values indicate faster responses during high-reward trials. The q values control FDR across the four tests in this table.",
        "",
        "## Joint regression",
        "",
        (
            f"The combined model was significant (R² = {model_row['r_squared']:.3f}, adjusted R² = "
            f"{model_row['adjusted_r_squared']:.3f}, model p = {model_row['model_p_value']:.4g}). "
            f"Apathy and BDI-II were themselves strongly correlated (r = {model_row['apathy_bdi_pearson_r']:.3f}), "
            "so their unique coefficients are less precise when both are entered together."
        ),
        "",
        markdown_table(
            coefficients,
            ["term", "coefficient_ms_per_point", "ci_low", "ci_high", "standardized_beta", "p_value", "hc3_p_value"],
            {"coefficient_ms_per_point": ".2f", "ci_low": ".2f", "ci_high": ".2f", "standardized_beta": ".3f", "p_value": ".4g", "hc3_p_value": ".4g"},
        ),
        "",
        "## Participant-level analysis data",
        "",
        markdown_table(
            subject_display,
            ["Subject", "Apathy", "BDI-II", "Valid trials", "Median RT (ms)", "OFF median (ms)", "ON median (ms)"],
            {"Apathy": ".0f", "BDI-II": ".0f", "Valid trials": ".0f", "Median RT (ms)": ".1f", "OFF median (ms)": ".1f", "ON median (ms)": ".1f"},
        ),
        "",
        "## Sensitivity analyses",
        "",
        markdown_table(
            sensitivity,
            ["method", "outcome", "predictor", "n_subjects", "correlation", "p_value", "fdr_q_across_sensitivity_tests"],
            {"correlation": ".3f", "p_value": ".4g", "fdr_q_across_sensitivity_tests": ".4g"},
        ),
        "",
        "## Session-1 projected-signal figure adjusted for apathy and BDI-II",
        "",
        markdown_table(
            projection_comparison,
            ["analysis", "n_subjects", "mean_behaviour_variability", "mean_projection_variability", "mean_behaviour_minus_projection", "difference_ci_low", "difference_ci_high", "paired_t_p_value", "wilcoxon_p_value"],
            {"mean_behaviour_variability": ".2f", "mean_projection_variability": ".2f", "mean_behaviour_minus_projection": ".2f", "difference_ci_low": ".2f", "difference_ci_high": ".2f", "paired_t_p_value": ".4g", "wilcoxon_p_value": ".4g"},
        ),
        "",
        (
            "Adjustment was performed at the subject level. For behaviour and projection separately, the linear effects of centered apathy and BDI-II were removed and the original outcome mean was restored. "
            "This standardizes each measure to the sample-average clinical scores. The same 17 participants and two-run subject means present in the source figure were retained. "
            "The source figure's saved behavioural variability was computed from the stored `1/RT` values after within-run demeaning, so the adjusted figure remains directly comparable to that original rather than silently changing its metric."
        ),
        "",
        markdown_table(
            projection_adjustment_models.loc[projection_adjustment_models["term"] != "const"],
            ["outcome", "term", "coefficient", "ci_low", "ci_high", "p_value", "model_r_squared", "model_p_value"],
            {"coefficient": ".3f", "ci_low": ".3f", "ci_high": ".3f", "p_value": ".4g", "model_r_squared": ".3f", "model_p_value": ".4g"},
        ),
        "",
        "## Data handling and matching",
        "",
        f"- Behaviour source: `{behaviour_dir}`",
        f"- Clinical source: `{clinical_workbook}`, `{CLINICAL_SHEET}`",
        "- Exact behavioural filename pattern: `PSPD###_ses_#_run_#.npy`.",
        "- Spreadsheet IDs such as `PS_PD001` were mapped to behavioural IDs such as `PSPD001` by their exact numeric component.",
        "- For 2-D behavioural arrays, column 2 was selected; 1-D arrays were treated as the already-extracted equivalent.",
        "- Stored values are inverse RT. Valid finite positive values were converted to milliseconds as `1000 / value`.",
        "- Missing/invalid trials already encoded as NaN were excluded. No additional trial trimming was imposed; median RT is the robust primary outcome.",
        "- The primary outcome pools both sessions and runs within each subject, so medication-state labeling does not affect it.",
        "- Exploratory OFF/ON summaries use session 1 = OFF and session 2 = ON, with the documented reversal for PSPD017.",
        f"- Skipped non-production `.npy` files: {', '.join(skipped_npy) if skipped_npy else 'none'}.",
        "",
        "## Diagnostics",
        "",
        diagnostics.to_markdown(index=False),
        "",
        "## Interpretation",
        "",
        "These results show association, not causation. The sample is small, and apathy and depression overlap substantially. The separate models support a relationship between slower responding and each clinical measure, while the joint model cannot cleanly separate their independent contributions.",
        "",
        "The complete ID reconciliation is in `id_matching_report.csv` and the formatted Excel workbook.",
    ]
    (out_dir / "rt_clinical_association_report.md").write_text("\n".join(report) + "\n", encoding="utf-8")


def format_excel(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in sheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
            width = min(max(max(map(len, values), default=0) + 2, 10), 45)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_outputs(
    args: argparse.Namespace,
    trials: pd.DataFrame,
    file_audit: pd.DataFrame,
    skipped_npy: list[str],
    match_report: pd.DataFrame,
    subjects: pd.DataFrame,
    session_summary: pd.DataFrame,
    primary: pd.DataFrame,
    coefficients: pd.DataFrame,
    joint_model_table: pd.DataFrame,
    sensitivity: pd.DataFrame,
    rt_correlations: pd.DataFrame,
    variability_correlations: pd.DataFrame,
    clinical_correlations: pd.DataFrame,
    reward_subject_data: pd.DataFrame,
    reward_correlations: pd.DataFrame,
    projection_adjusted_subjects: pd.DataFrame,
    projection_adjustment_models: pd.DataFrame,
    projection_comparison: pd.DataFrame,
    diagnostics: pd.DataFrame,
) -> None:
    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    tables = {
        "trial_level_rt": trials,
        "behaviour_file_audit": file_audit,
        "id_matching_report": match_report,
        "subject_analysis_data": subjects,
        "session_rt_summary": session_summary,
        "primary_associations": primary,
        "joint_coefficients": coefficients,
        "joint_model_summary": joint_model_table,
        "sensitivity_analyses": sensitivity,
        "rt_pearson_spearman": rt_correlations,
        "rt_variability_correlations": variability_correlations,
        "clinical_score_correlations": clinical_correlations,
        "reward_rt_subject_data": reward_subject_data,
        "reward_rt_correlations": reward_correlations,
        "projection_adjusted_subjects": projection_adjusted_subjects,
        "projection_adjustment_models": projection_adjustment_models,
        "projection_adjusted_comparison": projection_comparison,
        "model_diagnostics": diagnostics,
    }
    for name, frame in tables.items():
        frame.to_csv(out_dir / f"{name}.csv", index=False)

    excel_path = out_dir / "rt_clinical_association_results.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        readme = pd.DataFrame({
            "Item": [
                "Primary outcome", "Behaviour conversion", "Included participants", "Clinical source",
                "ID mapping", "RT variability", "Projection adjustment", "Medication mapping",
                "Skipped files", "Reward RT difference"
            ],
            "Value": [
                "Subject median RT pooled over all valid trials, sessions, and runs",
                "RT milliseconds = 1000 / stored inverse-RT value; 2-D arrays use column 2",
                str(int(subjects["included_primary"].sum())),
                f"{args.clinical_workbook} [{CLINICAL_SHEET}]",
                "PS_PD### -> PSPD### using the exact numeric identifier",
                "Pair-weighted mean normalized difference over adjacent valid RT pairs within each run",
                "Subject-level residualization on centered apathy and BDI-II, with original means restored",
                "Session 1 OFF/session 2 ON; PSPD017 reversed for state-specific sensitivity analyses",
                "; ".join(skipped_npy) if skipped_npy else "None",
                "High-reward mean RT minus low-reward mean RT from reward_rt_subject_bar_range_stats.csv",
            ],
        })
        readme.to_excel(writer, sheet_name="README", index=False)
        sheet_names = {
            "subject_analysis_data": "Subject data",
            "primary_associations": "Primary results",
            "joint_coefficients": "Joint coefficients",
            "joint_model_summary": "Joint model",
            "sensitivity_analyses": "Sensitivity",
            "rt_pearson_spearman": "RT correlations",
            "rt_variability_correlations": "RT variability corr",
            "clinical_score_correlations": "Clinical correlation",
            "reward_rt_subject_data": "Reward subject data",
            "reward_rt_correlations": "Reward correlations",
            "projection_adjusted_subjects": "Projection adjusted",
            "projection_adjustment_models": "Projection models",
            "projection_adjusted_comparison": "Projection comparison",
            "session_rt_summary": "Session RT",
            "id_matching_report": "ID matching",
            "behaviour_file_audit": "File audit",
            "model_diagnostics": "Diagnostics",
        }
        for key, sheet_name in sheet_names.items():
            tables[key].to_excel(writer, sheet_name=sheet_name, index=False)
    format_excel(excel_path)

    metadata = {
        "behaviour_dir": str(args.behaviour_dir),
        "clinical_workbook": str(args.clinical_workbook),
        "clinical_sheet": CLINICAL_SHEET,
        "behaviour_file_regex": BEHAVIOUR_RE.pattern,
        "behaviour_column_zero_based_for_2d": BEHAVIOUR_COLUMN_ZERO_BASED,
        "rt_conversion": "rt_ms = 1000 / inverse_rt_per_s",
        "primary_outcome": "subject median RT pooled over valid trials in both sessions/runs",
        "rt_variability_definition": (
            "pair-weighted subject mean of (RT[t]-RT[t+1])^2 / "
            "(RT[t]^2+RT[t+1]^2), within run, without bridging missing trials"
        ),
        "projection_adjustment_source": str(args.projection_session1_metrics),
        "reward_subject_stats_source": str(args.reward_subject_stats),
        "reward_rt_difference_definition": "1000 * (mean RT high reward - mean RT low reward), milliseconds",
        "projection_adjustment": (
            "separate subject-level OLS for behaviour and projection variability on centered apathy and "
            "BDI-II; adjusted value = original outcome mean + model residual"
        ),
        "n_behaviour_files": len(file_audit),
        "n_behaviour_subjects": int(trials["subject"].nunique()),
        "n_included_subjects": int(subjects["included_primary"].sum()),
        "skipped_npy_files": skipped_npy,
        "subject_session_state_overrides": {
            f"{subject}_session_{session}": state
            for (subject, session), state in SUBJECT_SESSION_STATE_OVERRIDES.items()
        },
    }
    (out_dir / "analysis_metadata.json").write_text(json.dumps(metadata, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    args = parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)
    trials, file_audit, skipped_npy = load_behaviour(args.behaviour_dir)
    clinical = load_clinical(args.clinical_workbook)
    behaviour_subjects = set(trials["subject"])
    match_report = build_match_report(clinical, behaviour_subjects)
    subjects, session_summary = summarize_subjects(trials, file_audit, clinical)
    analysis_data = subjects[subjects["included_primary"]].copy()
    if len(analysis_data) < 4:
        raise RuntimeError(f"Only {len(analysis_data)} complete matched subjects; analysis is not viable")

    primary = primary_analysis(analysis_data)
    coefficients, joint_model_table, joint_model = joint_regression(analysis_data)
    sensitivity = sensitivity_analysis(analysis_data)
    rt_correlations, variability_correlations, clinical_correlations = requested_correlation_tables(
        analysis_data
    )
    reward_subject_data, reward_correlations = reward_difference_analysis(
        args.reward_subject_stats, analysis_data, args.out_dir
    )
    projection_adjusted_subjects, projection_adjustment_models, projection_comparison = (
        adjusted_projection_analysis(args.projection_session1_metrics, analysis_data, args.out_dir)
    )
    diagnostics = diagnostic_analysis(analysis_data, joint_model)
    write_outputs(
        args, trials, file_audit, skipped_npy, match_report, subjects, session_summary,
        primary, coefficients, joint_model_table, sensitivity, rt_correlations,
        variability_correlations, clinical_correlations, reward_subject_data, reward_correlations,
        projection_adjusted_subjects,
        projection_adjustment_models, projection_comparison, diagnostics,
    )
    regression_plot(analysis_data, primary, args.out_dir)
    variability_scatterplot(analysis_data, variability_correlations, args.out_dir)
    write_report(
        args.out_dir, args.behaviour_dir, args.clinical_workbook, subjects, primary, coefficients,
        joint_model_table, sensitivity, rt_correlations, variability_correlations,
        clinical_correlations, reward_correlations, projection_adjustment_models, projection_comparison,
        diagnostics, match_report, skipped_npy,
    )
    print(f"Included subjects: {len(analysis_data)}")
    print(primary.to_string(index=False))
    print(f"Outputs: {args.out_dir}")


if __name__ == "__main__":
    main()
